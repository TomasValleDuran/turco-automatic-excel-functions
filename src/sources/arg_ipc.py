"""Download INDEC apendice4.xlsx and read tab '4.1.1 IPC NG'.

The ARG IPC sheet in Macro.xlsx is a 1:1 mirror of this tab. We return the
ordered list of data rows: (periodo, nacional, gba, pampeana, nea, noa, cuyo,
patagonia). Header/footnote rows are skipped (a data row has a period label in
col 0 and numeric values).

Note: the economia.gob.ar host serves a certificate chain that some clients
reject, so TLS verification is relaxed for this download only.
"""
from __future__ import annotations

import io
import warnings

import requests
import urllib3
from openpyxl import load_workbook

URL = "https://www.economia.gob.ar/download/infoeco/apendice4.xlsx"
SHEET = "4.1.1 IPC NG"
HEADERS = {"User-Agent": "Mozilla/5.0 (macro-updater)"}


def _is_data_row(row: tuple) -> bool:
    label, *vals = row[:8]
    if label is None and all(v is None for v in vals):
        return False
    numeric = [v for v in vals if isinstance(v, (int, float))]
    return len(numeric) >= 1


def fetch_ipc() -> list[tuple]:
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        resp = requests.get(URL, headers=HEADERS, timeout=120, verify=False)
    resp.raise_for_status()

    wb = load_workbook(io.BytesIO(resp.content), data_only=True, read_only=True)
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))

    # find the header row ('Período' in col 0), data starts 2 rows below
    # (there is a code row between header and data).
    start = None
    for i, r in enumerate(rows):
        if r and str(r[0]).strip().lower() == "período":
            start = i + 2
            break
    if start is None:
        raise ValueError("Could not find 'Período' header in apendice4")

    data: list[tuple] = []
    for r in rows[start:]:
        if str(r[0]).strip().lower().startswith("fuente"):
            break
        if _is_data_row(r):
            data.append(tuple(r[:8]))
    return data


if __name__ == "__main__":
    d = fetch_ipc()
    print(f"{len(d)} periods; last 3:")
    for r in d[-3:]:
        print(r)
