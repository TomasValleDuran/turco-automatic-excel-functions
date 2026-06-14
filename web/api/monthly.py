"""Monthly values for copy-paste: USA CPI + ARG IPC (latest period).

CPI is a quick HTML scrape. IPC downloads INDEC's apendice4.xlsx (~11 MB) and
reads the '4.1.1 IPC NG' tab, so this endpoint is slower (see maxDuration in
vercel.json). Each source fails independently so one slow/broken source never
hides the other.
"""
from http.server import BaseHTTPRequestHandler
import datetime as dt
import io
import json

import requests
import urllib3
from bs4 import BeautifulSoup
from openpyxl import load_workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
UA = {"User-Agent": "Mozilla/5.0 (macro-tool)"}
REGIONS = ["NACIONAL", "GBA", "PAMPEANA", "NEA", "NOA", "CUYO", "PATAGONIA"]


def fetch_cpi() -> dict:
    url = ("https://www.usinflationcalculator.com/inflation/"
           "consumer-price-index-and-annual-percent-changes-from-1913-to-2008/")
    r = requests.get(url, headers=UA, timeout=20)
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "lxml")
    latest = None
    for row in soup.find_all("tr"):
        cells = [c.get_text(strip=True) for c in row.find_all("td")]
        if not cells:
            continue
        try:
            y = int(cells[0])
        except ValueError:
            continue
        if not (1900 <= y <= 2100):
            continue
        for m in range(12, 0, -1):
            if m < len(cells):
                try:
                    v = float(cells[m].replace(",", ""))
                except ValueError:
                    continue
                if latest is None or (y, m) > (latest["year"], latest["month"]):
                    latest = {"year": y, "month": m, "value": v}
                break
    if not latest:
        raise ValueError("Sin datos CPI")
    return latest


def fetch_ipc() -> dict:
    url = "https://www.economia.gob.ar/download/infoeco/apendice4.xlsx"
    r = requests.get(url, headers=UA, timeout=50, verify=False)
    r.raise_for_status()
    wb = load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
    ws = wb["4.1.1 IPC NG"]
    rows = list(ws.iter_rows(values_only=True))
    start = None
    for i, rr in enumerate(rows):
        if rr and str(rr[0]).strip().lower() == "período":
            start = i + 2
            break
    # The INDEC file returns cached formula values inconsistently (stray None /
    # '#VALUE!' rows). Skip anything that isn't a complete region row (7 numbers)
    # and stop only at the 'Fuente:' footnote — robust across downloads.
    data = []
    for rr in rows[start:]:
        if str(rr[0]).strip().lower().startswith("fuente"):
            break
        nums = [c for c in rr[1:8] if isinstance(c, (int, float))]
        if len(nums) >= 7:
            data.append(rr[:8])
    if not data:
        raise ValueError("Sin datos IPC")
    last = data[-1]
    per = last[0]
    if isinstance(per, dt.datetime):
        per = per.date().isoformat()
    elif isinstance(per, dt.date):
        per = per.isoformat()
    valores = [round(x, 4) if isinstance(x, (int, float)) else x for x in last[1:8]]
    return {"periodo": str(per), "columnas": REGIONS, "valores": valores}


class handler(BaseHTTPRequestHandler):
    def do_GET(self):
        out = {}
        for key, fn in (("cpi", fetch_cpi), ("ipc", fetch_ipc)):
            try:
                out[key] = fn()
            except Exception as e:  # noqa: BLE001
                out[key] = {"error": str(e)}
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(json.dumps(out).encode())
